"""Excel import service."""

from __future__ import annotations

import re
import uuid
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from app.extensions import db
from app.models import (
    Contract,
    EducationStatus,
    Employment,
    EmploymentStatus,
    GradeCatalog,
    ImportJob,
    ImportRow,
    ImportStatus,
    Passport,
    Person,
)
from app.services.employees import (
    create_person_with_employment,
    get_active_contract,
    get_active_passport,
    get_current_grade,
    get_current_name,
    get_current_position,
    sync_active_contract,
    sync_actual_grade,
    sync_employment_periods_from_import,
    sync_passport,
    update_position,
)
from app.services.grades import resolve_unknown_education_snapshot
from app.services.rule_engine import run_rule_engine
from app.services.tenure import (
    MAX_EMPLOYMENT_PERIODS,
    auto_mark_reached_awards,
    count_employment_periods,
    ensure_tenure_awards,
    employment_periods,
)
from app.utils.dates import (
    calculate_contract_start,
    format_display_date_ru,
    normalize_full_name,
    parse_flexible_date,
)

COLUMN_MAP = {
    "фио": "full_name",
    "full_name": "full_name",
    "должность": "title",
    "title": "title",
    "грейдов по должности": "position_grade",
    "грейд по должности": "position_grade",
    "position_grade": "position_grade",
    "фактический грейдов": "actual_grade",
    "фактический грейд": "actual_grade",
    "actual_grade": "actual_grade",
    "вуз": "education_status",
    "has_university": "education_status",
    "education_status": "education_status",
    "срок договора": "contract_term_years",
    "срок договора (лет)": "contract_term_years",
    "срок договора, лет": "contract_term_years",
    "срок контракта": "contract_term_years",
    "срок контракта (лет)": "contract_term_years",
    "contract_term_years": "contract_term_years",
    "окончание договора": "contract_end",
    "contract_end": "contract_end",
    "дата получения текущего грейда": "grade_date",
    "дата получения текущего грейды": "grade_date",
    "grade_date": "grade_date",
    "начало работы": "hire_date",
    "hire_date": "hire_date",
    "начало работы 1": "period_1_hire",
    "увольнение 1": "period_1_dismissal",
    "начало работы 2": "period_2_hire",
    "увольнение 2": "period_2_dismissal",
    "начало работы 3": "period_3_hire",
    "увольнение 3": "period_3_dismissal",
    "period_1_hire": "period_1_hire",
    "period_1_dismissal": "period_1_dismissal",
    "period_2_hire": "period_2_hire",
    "period_2_dismissal": "period_2_dismissal",
    "period_3_hire": "period_3_hire",
    "period_3_dismissal": "period_3_dismissal",
    "срок окончания паспорта": "passport_until",
    "passport_until": "passport_until",
    "№ п/п": "row_num",
    "row_num": "row_num",
}

DATE_FIELDS = {
    "hire_date",
    "contract_end",
    "grade_date",
    "passport_until",
    "period_1_hire",
    "period_1_dismissal",
    "period_2_hire",
    "period_2_dismissal",
    "period_3_hire",
    "period_3_dismissal",
}
PERIOD_FIELD_GROUPS = (
    ("period_1_hire", "period_1_dismissal"),
    ("period_2_hire", "period_2_dismissal"),
    ("period_3_hire", "period_3_dismissal"),
)
GRADE_FIELDS = ("position_grade", "actual_grade")
UNKNOWN_GRADE_WARNING_PREFIX = "Грейд «"
UNKNOWN_GRADE_WARNING_SUFFIX = "» не найден в справочнике"


def _normalize_header(value: str) -> str:
    return re.sub(r"\s+", " ", value.strip().lower())


def _parse_bool(value: Any) -> bool:
    parsed = _parse_bool_optional(value)
    return bool(parsed)


def _parse_bool_optional(value: Any) -> bool | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    text = str(value).strip().lower()
    if not text:
        return None
    if text in {"да", "yes", "1", "true", "+", "есть"}:
        return True
    if text in {"нет", "no", "0", "false", "-"}:
        return False
    return None


def _parse_education_status(value: Any) -> str | None:
    parsed = _parse_bool_optional(value)
    if parsed is True:
        return EducationStatus.YES.value
    if parsed is False:
        return EducationStatus.NO.value
    return None


def _parse_date(value: Any) -> date | None:
    return parse_flexible_date(value)


def _period_columns_used(data: dict[str, Any]) -> bool:
    return any(data.get(field) not in (None, "") for field, _ in PERIOD_FIELD_GROUPS)


def parse_employment_periods(data: dict[str, Any]) -> list[tuple[date, date | None]] | None:
    """Return imported employment periods or ``None`` when legacy hire_date mode is used."""
    if not _period_columns_used(data):
        return None

    periods: list[tuple[date, date | None]] = []
    for hire_field, dismiss_field in PERIOD_FIELD_GROUPS:
        hire = _parse_date(data.get(hire_field))
        dismiss = _parse_date(data.get(dismiss_field))
        if hire is None and dismiss is None:
            break
        periods.append((hire, dismiss))
    return periods


def validate_employment_periods(
    data: dict[str, Any],
    *,
    existing_period_count: int | None = None,
) -> list[str]:
    errors: list[str] = []
    periods = parse_employment_periods(data)
    if periods is None:
        return errors

    if not periods or periods[0][0] is None:
        errors.append("Для периода 1 укажите дату начала работы")
        return errors

    if len(periods) > MAX_EMPLOYMENT_PERIODS:
        errors.append("Можно указать не более 3 периодов работы")

    previous_dismissal: date | None = None
    for index, (hire, dismiss) in enumerate(periods, start=1):
        if hire is None:
            errors.append(f"Для периода {index} укажите дату начала работы")
            continue
        if dismiss is not None and dismiss < hire:
            errors.append(f"Увольнение периода {index} не может быть раньше начала")
        if index < len(periods) and dismiss is None:
            errors.append(
                f"Укажите дату увольнения для периода {index} перед следующим периодом"
            )
        if previous_dismissal is not None and hire <= previous_dismissal:
            errors.append(
                f"Начало периода {index} должно быть позже увольнения предыдущего периода"
            )
        previous_dismissal = dismiss

    legacy_hire = _parse_date(data.get("hire_date"))
    last_hire = periods[-1][0]
    if legacy_hire and last_hire and legacy_hire != last_hire:
        errors.append(
            "Колонка «Начало работы» должна совпадать с началом последнего периода"
        )

    if existing_period_count is not None and len(periods) < existing_period_count:
        errors.append(
            "Импорт не удаляет ранее заведённые периоды работы — укажите все периоды"
        )

    return errors


def effective_hire_date(data: dict[str, Any]) -> date | None:
    periods = parse_employment_periods(data)
    if periods:
        last_hire = periods[-1][0]
        return last_hire
    return _parse_date(data.get("hire_date"))


def _normalize_grade_name(name: str | None) -> str:
    if name is None:
        return ""
    return str(name).strip().lower()


def _resolve_grade_id(name: str | None) -> int | None:
    needle = _normalize_grade_name(name)
    if not needle:
        return None
    for grade in GradeCatalog.query.all():
        if grade.name.strip().lower() == needle:
            return grade.id
    return None


def _catalog_name_set() -> set[str]:
    names: set[str] = set()
    for grade in GradeCatalog.query.all():
        normalized = _normalize_grade_name(grade.name)
        if normalized:
            names.add(normalized)
    return names


def _row_grade_values(data: dict[str, Any]) -> list[str]:
    values: list[str] = []
    seen: set[str] = set()
    for key in GRADE_FIELDS:
        raw = data.get(key)
        if raw is None:
            continue
        text = str(raw).strip()
        needle = _normalize_grade_name(text)
        if not needle or needle in seen:
            continue
        seen.add(needle)
        values.append(text)
    return values


def _unknown_grade_warning(name: str) -> str:
    return f"{UNKNOWN_GRADE_WARNING_PREFIX}{name}{UNKNOWN_GRADE_WARNING_SUFFIX}"


def _is_unknown_grade_warning(message: str) -> bool:
    return message.startswith(UNKNOWN_GRADE_WARNING_PREFIX) and message.endswith(
        UNKNOWN_GRADE_WARNING_SUFFIX
    )


def _parse_contract_term_years(value: Any) -> float | None:
    if value is None or str(value).strip() == "":
        return None
    try:
        term = float(str(value).replace(",", ".").strip())
    except ValueError:
        return None
    if term <= 0:
        return None
    return term


def _contract_fields_present(data: dict[str, Any]) -> tuple[bool, bool]:
    has_end = data.get("contract_end") not in (None, "") and str(data.get("contract_end")).strip()
    has_term = data.get("contract_term_years") not in (None, "") and str(
        data.get("contract_term_years")
    ).strip()
    return bool(has_end), bool(has_term)


def _serialize_raw_data(data: dict[str, Any]) -> dict[str, Any]:
    serialized: dict[str, Any] = {}
    for key, value in data.items():
        if key.startswith("_"):
            continue
        if key in DATE_FIELDS:
            parsed = _parse_date(value)
            serialized[key] = parsed.isoformat() if parsed else None
            continue
        if key == "education_status":
            status = _parse_education_status(value)
            if status is None:
                serialized[key] = None
            else:
                serialized[key] = (
                    "Да" if status == EducationStatus.YES.value else "Нет"
                )
            continue
        if key == "contract_term_years":
            parsed = _parse_contract_term_years(value)
            serialized[key] = parsed
            continue
        serialized[key] = None if value is None else str(value)
    return serialized


def candidate_payload(person: Person) -> dict[str, str | None]:
    employment = (
        Employment.query.filter_by(person_id=person.id)
        .order_by(Employment.hire_date.desc())
        .first()
    )
    title = None
    if employment:
        position = get_current_position(employment)
        title = position.title if position else None
    return {
        "uuid": str(person.uuid),
        "full_name": get_current_name(person),
        "title": title,
    }


def parse_workbook(path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [_normalize_header(str(h or "")) for h in rows[0]]
    mapped_headers = [COLUMN_MAP.get(h, h) for h in headers]
    parsed: list[dict[str, Any]] = []

    for idx, row in enumerate(rows[1:], start=2):
        if not any(row):
            continue
        data = {mapped_headers[i]: row[i] for i in range(len(mapped_headers)) if i < len(row)}
        data.pop("uuid", None)
        data["_row_number"] = idx
        parsed.append(data)
    return parsed


def validate_row(data: dict[str, Any]) -> tuple[list[str], list[str]]:
    errors: list[str] = []
    warnings: list[str] = []

    if not data.get("full_name"):
        errors.append("Не указано ФИО")

    if not effective_hire_date(data):
        errors.append("Не указана или некорректна дата начала работы")

    errors.extend(validate_employment_periods(data))

    education_raw = data.get("education_status")
    if (
        education_raw is not None
        and str(education_raw).strip()
        and _parse_education_status(education_raw) is None
    ):
        errors.append("Поле ВУЗ должно содержать «Да» или «Нет»")

    has_end, has_term = _contract_fields_present(data)
    if has_end != has_term:
        errors.append("Укажите срок договора и дату окончания")
    elif has_end:
        contract_end = _parse_date(data.get("contract_end"))
        term_years = _parse_contract_term_years(data.get("contract_term_years"))
        if contract_end is None:
            errors.append("Некорректная дата окончания договора")
        if term_years is None:
            errors.append("Некорректный срок договора")
        if contract_end and term_years:
            try:
                calculate_contract_start(contract_end, term_years)
            except ValueError as exc:
                errors.append(str(exc))

    return errors, warnings


def find_person_candidates(full_name: str) -> list[Person]:
    normalized = normalize_full_name(str(full_name))
    matches: list[Person] = []
    for person in Person.query.all():
        current_name = get_current_name(person)
        if current_name and normalize_full_name(current_name) == normalized:
            matches.append(person)
    return matches


def dry_run_import(job: ImportJob, rows: list[dict[str, Any]]) -> None:
    summary = {"create": 0, "update": 0, "ambiguous": 0, "error": 0}
    for data in rows:
        row_number = data.get("_row_number", 0)
        errors, warnings = validate_row(data)
        action = None
        person_uuid = None
        candidates: list[dict[str, str | None]] | None = None

        if errors:
            action = "error"
            summary["error"] += 1
        else:
            matches = find_person_candidates(str(data.get("full_name", "")))
            if len(matches) == 1:
                action = "update"
                person_uuid = matches[0].uuid
                summary["update"] += 1
            elif len(matches) > 1:
                action = "ambiguous"
                candidates = [candidate_payload(person) for person in matches]
                summary["ambiguous"] += 1
                warnings.append("Найдено несколько кандидатов — выберите действие")
            else:
                action = "create"
                if _parse_education_status(data.get("education_status")) is None:
                    errors.append("Для нового сотрудника укажите ВУЗ: «Да» или «Нет»")
                    action = "error"
                    summary["error"] += 1
                else:
                    summary["create"] += 1

        db.session.add(
            ImportRow(
                import_job_id=job.id,
                row_number=row_number,
                raw_data=_serialize_raw_data(data),
                action=action,
                person_uuid=person_uuid,
                candidates=candidates,
                errors=errors or None,
                warnings=warnings or None,
            )
        )

    db.session.flush()
    job.status = ImportStatus.VALIDATED.value
    job.summary = summary
    annotate_unknown_grades(job)


def annotate_unknown_grades(job: ImportJob) -> None:
    catalog = _catalog_name_set()
    collected: dict[str, dict[str, Any]] = {}

    for row in job.rows:
        preserved = [item for item in (row.warnings or []) if not _is_unknown_grade_warning(item)]
        if row.errors:
            row.warnings = preserved or None
            continue

        unknown_in_row: list[str] = []
        for name in _row_grade_values(row.raw_data or {}):
            needle = _normalize_grade_name(name)
            if needle in catalog:
                continue
            unknown_in_row.append(name)
            entry = collected.setdefault(needle, {"name": name, "count": 0})
            entry["count"] += 1

        warnings = preserved + [_unknown_grade_warning(name) for name in unknown_in_row]
        row.warnings = warnings or None

    summary = dict(job.summary or {})
    summary["unknown_grades"] = list(collected.values())
    job.summary = summary


def _resolve_row_action(
    row: ImportRow,
    row_actions: dict[int, str | None],
) -> tuple[str | None, uuid.UUID | None]:
    chosen = row_actions.get(row.id)
    if chosen is None or chosen == "":
        return row.action, row.person_uuid

    chosen = str(chosen).strip()
    if chosen == "skip":
        return "skip", None
    if chosen == "create":
        return "create", None
    if chosen == "update":
        return "update", row.person_uuid
    if chosen.startswith("update:"):
        raw_uuid = chosen.split(":", 1)[1].strip()
        return "update", uuid.UUID(raw_uuid)
    return chosen, row.person_uuid


def _mark_row_result(row: ImportRow, result: str, message: str | None = None) -> None:
    row.result = result
    row.result_message = message


def _upsert_contract(
    employment: Employment,
    contract_end: date,
    term_years: float,
) -> None:
    start_date = calculate_contract_start(contract_end, term_years)
    existing_contract = Contract.query.filter_by(
        employment_id=employment.id,
        end_date=contract_end,
    ).first()
    if not existing_contract:
        db.session.add(
            Contract(
                employment_id=employment.id,
                start_date=start_date,
                end_date=contract_end,
                term_years=term_years,
                is_active=True,
            )
        )


def _upsert_passport(person: Person, passport_until: date) -> None:
    existing = Passport.query.filter_by(
        person_id=person.id,
        valid_until=passport_until,
        is_active=True,
    ).first()
    if existing:
        return
    db.session.add(
        Passport(
            person_id=person.id,
            valid_until=passport_until,
            is_active=True,
        )
    )


def confirm_import(
    job: ImportJob,
    row_actions: dict[int, str | None] | None = None,
    *,
    mark_reached_tenure: bool = True,
    update_existing_tenure: bool = True,
) -> None:
    row_actions = row_actions or {}
    report = {
        "created": 0,
        "updated": 0,
        "skipped": 0,
        "errors": 0,
        "tenure_marked": 0,
        "skipped_reasons": {},
    }

    def bump_skip(reason: str) -> None:
        report["skipped"] += 1
        reasons = report["skipped_reasons"]
        reasons[reason] = reasons.get(reason, 0) + 1

    try:
        for row in job.rows:
            if row.action == "error":
                report["errors"] += 1
                _mark_row_result(row, "error", "Строка содержит ошибки валидации")
                continue

            action, person_uuid = _resolve_row_action(row, row_actions)

            if action == "ambiguous":
                bump_skip("ambiguous_unresolved")
                _mark_row_result(row, "skipped", "Не выбрано действие для дубликата")
                continue

            if action == "skip":
                bump_skip("skipped_by_user")
                _mark_row_result(row, "skipped", "Пропущено пользователем")
                continue

            data = row.raw_data or {}
            hire_date = effective_hire_date(data)
            if not hire_date:
                bump_skip("no_hire_date")
                _mark_row_result(row, "skipped", "Некорректная дата начала работы")
                continue

            periods = parse_employment_periods(data)
            person: Person | None = None
            employment: Employment | None = None
            created = False
            title = str(data.get("title") or "Не указана")
            position_grade_id = _resolve_grade_id(data.get("position_grade"))

            if action == "update":
                if not person_uuid:
                    bump_skip("no_person")
                    _mark_row_result(row, "skipped", "Не выбран сотрудник для обновления")
                    continue
                person = Person.query.filter_by(uuid=person_uuid).first()
                if not person:
                    bump_skip("person_not_found")
                    _mark_row_result(row, "skipped", "Сотрудник для обновления не найден")
                    continue
                if periods is not None:
                    period_errors = validate_employment_periods(
                        data,
                        existing_period_count=count_employment_periods(
                            person.id,
                            job.company_id,
                        ),
                    )
                    if period_errors:
                        bump_skip("invalid_employment_periods")
                        _mark_row_result(row, "skipped", period_errors[0])
                        continue
                    employment = sync_employment_periods_from_import(
                        person,
                        job.company_id,
                        periods,
                        title=title,
                        position_grade_id=position_grade_id,
                    )
                else:
                    employment = (
                        Employment.query.filter_by(
                            person_id=person.id,
                            company_id=job.company_id,
                        )
                        .order_by(Employment.hire_date.desc())
                        .first()
                    )
            elif action == "create":
                education_status = _parse_education_status(data.get("education_status"))
                if education_status is None:
                    bump_skip("missing_education_status")
                    _mark_row_result(
                        row,
                        "skipped",
                        "Для нового сотрудника укажите ВУЗ: «Да» или «Нет»",
                    )
                    continue
                initial_hire = periods[0][0] if periods else hire_date
                person, employment = create_person_with_employment(
                    company_id=job.company_id,
                    full_name=str(data.get("full_name", "")),
                    hire_date=initial_hire,
                    title=title,
                    position_grade_id=position_grade_id,
                    education_status=education_status,
                )
                row.person_uuid = person.uuid
                created = True
                if periods is not None:
                    employment = sync_employment_periods_from_import(
                        person,
                        job.company_id,
                        periods,
                        title=title,
                        position_grade_id=position_grade_id,
                    )
            else:
                bump_skip("unknown_action")
                _mark_row_result(row, "skipped", f"Неизвестное действие: {action}")
                continue

            if employment is None and person:
                employment = (
                    Employment.query.filter_by(person_id=person.id, company_id=job.company_id)
                    .order_by(Employment.hire_date.desc())
                    .first()
                )

            if employment is None or person is None:
                bump_skip("no_employment")
                _mark_row_result(row, "skipped", "Не найдено трудоустройство")
                continue

            education_status = _parse_education_status(data.get("education_status"))
            if education_status is not None:
                person.education_status = education_status
                resolve_unknown_education_snapshot(employment, education_status)

            if not created and periods is None:
                if hire_date != employment.hire_date:
                    employment.hire_date = hire_date
                if title.strip():
                    current_position = get_current_position(employment)
                    current_title = current_position.title if current_position else ""
                    if title != current_title:
                        update_position(
                            employment,
                            title,
                            position_grade_id,
                            hire_date,
                        )
            elif not created and periods is not None and title.strip():
                current_position = get_current_position(employment)
                current_title = current_position.title if current_position else ""
                if title != current_title:
                    update_position(
                        employment,
                        title,
                        position_grade_id,
                        employment.hire_date,
                    )

            contract_end = _parse_date(data.get("contract_end"))
            term_years = _parse_contract_term_years(data.get("contract_term_years"))

            if contract_end is not None or term_years is not None:
                sync_active_contract(
                    employment,
                    contract_end,
                    term_years=term_years,
                )

            grade_date = _parse_date(data.get("grade_date"))
            grade_id = _resolve_grade_id(data.get("actual_grade"))
            if grade_id and grade_date:
                sync_actual_grade(employment, grade_id, grade_date)

            passport_until = _parse_date(data.get("passport_until"))
            if passport_until:
                sync_passport(person, passport_until)

            awards = ensure_tenure_awards(person.id, employment.company_id)
            if mark_reached_tenure and (created or update_existing_tenure):
                report["tenure_marked"] += auto_mark_reached_awards(awards)

            if created:
                report["created"] += 1
                _mark_row_result(row, "created")
            else:
                report["updated"] += 1
                _mark_row_result(row, "updated")

        preview = {
            key: (job.summary or {}).get(key, 0)
            for key in ("create", "update", "ambiguous", "error")
        }
        job.summary = {**preview, **report}
        job.status = ImportStatus.CONFIRMED.value
        job.error_message = None
        db.session.flush()
        run_rule_engine(job.company_id)
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        job.status = ImportStatus.FAILED.value
        job.error_message = str(exc)
        db.session.commit()
        raise


def export_template(company_id: int, path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Employees"
    headers = [
        "№ п/п",
        "ФИО",
        "Должность",
        "Грейд по должности",
        "Фактический грейд",
        "ВУЗ",
        "Срок договора (лет)",
        "Окончание договора",
        "Дата получения текущего грейда",
        "Начало работы 1",
        "Увольнение 1",
        "Начало работы 2",
        "Увольнение 2",
        "Начало работы 3",
        "Увольнение 3",
        "Начало работы",
        "Срок окончания паспорта",
    ]
    ws.append(headers)

    seen_person_ids: set[int] = set()
    row_num = 0
    employments = Employment.query.filter_by(company_id=company_id).all()
    for employment in employments:
        person = employment.person
        if person.id in seen_person_ids:
            continue
        seen_person_ids.add(person.id)
        row_num += 1

        periods = employment_periods(person.id, company_id)
        active = next(
            (item for item in reversed(periods) if item.status == EmploymentStatus.ACTIVE.value),
            periods[-1] if periods else employment,
        )
        position = get_current_position(active)
        contract = get_active_contract(active)
        grade = get_current_grade(active)
        passport = get_active_passport(person)

        period_cells: list[str] = []
        for index in range(MAX_EMPLOYMENT_PERIODS):
            if index < len(periods):
                period = periods[index]
                period_cells.append(format_display_date_ru(period.hire_date))
                period_cells.append(
                    format_display_date_ru(period.dismissal_date)
                    if period.dismissal_date
                    else ""
                )
            else:
                period_cells.extend(["", ""])

        ws.append(
            [
                row_num,
                get_current_name(person),
                position.title if position else "",
                position.position_grade.name if position and position.position_grade else "",
                grade.grade.name if grade else "",
                (
                    "Да"
                    if person.education_status == EducationStatus.YES.value
                    else "Нет"
                    if person.education_status == EducationStatus.NO.value
                    else "Неизвестно"
                ),
                contract.term_years if contract and contract.term_years is not None else "",
                format_display_date_ru(contract.end_date) if contract else "",
                format_display_date_ru(grade.assigned_date) if grade else "",
                *period_cells,
                format_display_date_ru(active.hire_date),
                format_display_date_ru(passport.valid_until) if passport else "",
            ]
        )

    wb.save(path)


# Backward-compatible alias for older callers/tests.
export_template_with_uuids = export_template
