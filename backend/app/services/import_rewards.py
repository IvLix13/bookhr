"""Excel import service for rewards (поощрения)."""

from __future__ import annotations

import re
import uuid
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from app.extensions import db
from app.models import Employment, ImportJob, ImportRow, ImportStatus, Person, Reward, RewardStatus
from app.services.employees import get_current_name
from app.services.import_excel import candidate_payload, find_person_candidates
from app.services.rewards import create_reward, update_reward
from app.utils.dates import format_display_date_ru, parse_flexible_date

COLUMN_MAP = {
    "фио": "full_name",
    "full_name": "full_name",
    "вид поощрения": "reward_type",
    "reward_type": "reward_type",
    "состояние": "status",
    "status": "status",
    "указание на вручение": "directive_text",
    "directive_text": "directive_text",
    "дата вручения": "delivered_date",
    "delivered_date": "delivered_date",
    "примечание": "notes",
    "notes": "notes",
    "№ п/п": "row_num",
    "row_num": "row_num",
}

DATE_FIELDS = {"delivered_date"}

STATUS_ALIASES = {
    "не вручено": RewardStatus.NOT_DELIVERED.value,
    "not_delivered": RewardStatus.NOT_DELIVERED.value,
    "в кадрах": RewardStatus.IN_HR.value,
    "in_hr": RewardStatus.IN_HR.value,
    "вручено": RewardStatus.DELIVERED.value,
    "delivered": RewardStatus.DELIVERED.value,
    "доп. статус 1": RewardStatus.EXTRA_1.value,
    "extra_1": RewardStatus.EXTRA_1.value,
    "доп. статус 2": RewardStatus.EXTRA_2.value,
    "extra_2": RewardStatus.EXTRA_2.value,
    "доп. статус 3": RewardStatus.EXTRA_3.value,
    "extra_3": RewardStatus.EXTRA_3.value,
}

STATUS_LABELS = {
    RewardStatus.NOT_DELIVERED.value: "Не вручено",
    RewardStatus.IN_HR.value: "В кадрах",
    RewardStatus.DELIVERED.value: "Вручено",
    RewardStatus.EXTRA_1.value: "Доп. статус 1",
    RewardStatus.EXTRA_2.value: "Доп. статус 2",
    RewardStatus.EXTRA_3.value: "Доп. статус 3",
}


def _normalize_header(value: str) -> str:
    return re.sub(r"\s+", " ", value.strip().lower())


def _parse_date(value: Any) -> date | None:
    return parse_flexible_date(value)


def _parse_status(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip().lower()
    if not text:
        return None
    return STATUS_ALIASES.get(text)


def _normalize_reward_type(value: Any) -> str:
    return str(value or "").strip()


def _serialize_raw_data(data: dict[str, Any]) -> dict[str, Any]:
    serialized: dict[str, Any] = {}
    for key, value in data.items():
        if key.startswith("_"):
            continue
        if key in DATE_FIELDS:
            parsed = _parse_date(value)
            serialized[key] = parsed.isoformat() if parsed else None
            continue
        if key == "status":
            parsed = _parse_status(value)
            if parsed is None and value is not None and str(value).strip():
                serialized[key] = str(value).strip()
            else:
                serialized[key] = parsed or RewardStatus.NOT_DELIVERED.value
            continue
        serialized[key] = None if value is None else str(value)
    return serialized


def parse_rewards_workbook(path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [_normalize_header(str(h or "")) for h in rows[0]]
    mapped_headers = [COLUMN_MAP.get(h, h) for h in headers]
    parsed: list[dict[str, Any]] = []

    for idx, row in enumerate(rows[1:], start=2):
        if not any(row):
            continue
        data = {mapped_headers[i]: row[i] for i in range(len(mapped_headers)) if i < len(row)}
        data.pop("uuid", None)
        data["_row_number"] = idx
        parsed.append(data)
    return parsed


def validate_reward_row(data: dict[str, Any]) -> tuple[list[str], list[str]]:
    errors: list[str] = []
    warnings: list[str] = []

    if not data.get("full_name"):
        errors.append("Не указано ФИО")

    reward_type = _normalize_reward_type(data.get("reward_type"))
    if not reward_type:
        errors.append("Не указан вид поощрения")

    raw_status = data.get("status")
    if raw_status is not None and str(raw_status).strip():
        if _parse_status(raw_status) is None:
            errors.append("Некорректное состояние поощрения")

    delivered_raw = data.get("delivered_date")
    if delivered_raw is not None and str(delivered_raw).strip():
        if _parse_date(delivered_raw) is None:
            errors.append("Некорректная дата вручения")

    return errors, warnings


def _find_existing_reward(employment_id: int, reward_type: str) -> Reward | None:
    needle = reward_type.strip().lower()
    rewards = (
        Reward.query.filter_by(employment_id=employment_id)
        .order_by(Reward.updated_at.desc())
        .all()
    )
    for reward in rewards:
        if reward.reward_type.strip().lower() == needle:
            return reward
    return None


def _employment_for_person(person: Person, company_id: int) -> Employment | None:
    return (
        Employment.query.filter_by(person_id=person.id, company_id=company_id)
        .order_by(Employment.hire_date.desc())
        .first()
    )


def dry_run_rewards_import(job: ImportJob, rows: list[dict[str, Any]]) -> None:
    summary = {"create": 0, "update": 0, "ambiguous": 0, "error": 0}
    for data in rows:
        row_number = data.get("_row_number", 0)
        errors, warnings = validate_reward_row(data)
        action = None
        person_uuid = None
        candidates: list[dict[str, str | None]] | None = None

        if errors:
            action = "error"
            summary["error"] += 1
        else:
            matches = find_person_candidates(str(data.get("full_name", "")))
            reward_type = _normalize_reward_type(data.get("reward_type"))
            if len(matches) == 1:
                person = matches[0]
                person_uuid = person.uuid
                employment = _employment_for_person(person, job.company_id)
                if employment is None:
                    action = "error"
                    errors.append("Не найдено трудоустройство сотрудника в компании")
                    summary["error"] += 1
                else:
                    existing = _find_existing_reward(employment.id, reward_type)
                    if existing:
                        action = "update"
                        summary["update"] += 1
                    else:
                        action = "create"
                        summary["create"] += 1
            elif len(matches) > 1:
                action = "ambiguous"
                candidates = [candidate_payload(person) for person in matches]
                summary["ambiguous"] += 1
                warnings.append("Найдено несколько кандидатов — выберите сотрудника")
            else:
                action = "error"
                errors.append("Сотрудник с таким ФИО не найден")
                summary["error"] += 1

        db.session.add(
            ImportRow(
                import_job_id=job.id,
                row_number=row_number,
                raw_data=_serialize_raw_data(data),
                action=action,
                person_uuid=person_uuid,
                candidates=candidates,
                errors=errors or None,
                warnings=warnings or None,
            )
        )

    job.status = ImportStatus.VALIDATED.value
    job.summary = summary


def _resolve_row_action(
    row: ImportRow,
    row_actions: dict[int, str | None],
) -> tuple[str | None, uuid.UUID | None]:
    chosen = row_actions.get(row.id)
    if chosen is None or chosen == "":
        return row.action, row.person_uuid

    chosen = str(chosen).strip()
    if chosen == "skip":
        return "skip", None
    if chosen == "update":
        return "update", row.person_uuid
    if chosen.startswith("update:"):
        raw_uuid = chosen.split(":", 1)[1].strip()
        return "select", uuid.UUID(raw_uuid)
    return chosen, row.person_uuid


def _mark_row_result(row: ImportRow, result: str, message: str | None = None) -> None:
    row.result = result
    row.result_message = message


def confirm_rewards_import(
    job: ImportJob,
    row_actions: dict[int, str | None] | None = None,
) -> None:
    row_actions = row_actions or {}
    report = {
        "created": 0,
        "updated": 0,
        "skipped": 0,
        "errors": 0,
        "skipped_reasons": {},
    }

    def bump_skip(reason: str) -> None:
        report["skipped"] += 1
        reasons = report["skipped_reasons"]
        reasons[reason] = reasons.get(reason, 0) + 1

    try:
        for row in job.rows:
            if row.action == "error":
                report["errors"] += 1
                _mark_row_result(row, "error", "Строка содержит ошибки валидации")
                continue

            action, person_uuid = _resolve_row_action(row, row_actions)

            if action == "ambiguous":
                bump_skip("ambiguous_unresolved")
                _mark_row_result(row, "skipped", "Не выбран сотрудник для дубликата")
                continue

            if action == "skip":
                bump_skip("skipped_by_user")
                _mark_row_result(row, "skipped", "Пропущено пользователем")
                continue

            if action not in {"create", "update", "select"}:
                bump_skip("unknown_action")
                _mark_row_result(row, "skipped", f"Неизвестное действие: {action}")
                continue

            if not person_uuid:
                bump_skip("no_person")
                _mark_row_result(row, "skipped", "Не выбран сотрудник")
                continue

            person = Person.query.filter_by(uuid=person_uuid).first()
            if not person:
                bump_skip("person_not_found")
                _mark_row_result(row, "skipped", "Сотрудник не найден")
                continue

            employment = _employment_for_person(person, job.company_id)
            if employment is None:
                bump_skip("no_employment")
                _mark_row_result(row, "skipped", "Не найдено трудоустройство")
                continue

            data = row.raw_data or {}
            reward_type = _normalize_reward_type(data.get("reward_type"))
            if not reward_type:
                bump_skip("no_reward_type")
                _mark_row_result(row, "skipped", "Не указан вид поощрения")
                continue

            status = _parse_status(data.get("status")) or RewardStatus.NOT_DELIVERED.value
            delivered_date = _parse_date(data.get("delivered_date"))
            directive_text = data.get("directive_text") or None
            notes = data.get("notes") or None
            if isinstance(directive_text, str):
                directive_text = directive_text.strip() or None
            if isinstance(notes, str):
                notes = notes.strip() or None

            existing = _find_existing_reward(employment.id, reward_type)
            row.person_uuid = person.uuid

            if existing:
                update_reward(
                    existing,
                    {
                        "reward_type": reward_type,
                        "status": status,
                        "directive_text": directive_text,
                        "delivered_date": (
                            delivered_date.isoformat() if delivered_date else None
                        ),
                        "notes": notes,
                    },
                )
                report["updated"] += 1
                _mark_row_result(row, "updated")
            else:
                create_reward(
                    employment_id=employment.id,
                    reward_type=reward_type,
                    status=status,
                    directive_text=directive_text,
                    delivered_date=delivered_date,
                    notes=notes,
                )
                report["created"] += 1
                _mark_row_result(row, "created")

        preview = {
            key: (job.summary or {}).get(key, 0)
            for key in ("create", "update", "ambiguous", "error")
        }
        job.summary = {**preview, **report}
        job.status = ImportStatus.CONFIRMED.value
        job.error_message = None
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        job.status = ImportStatus.FAILED.value
        job.error_message = str(exc)
        db.session.commit()
        raise


def export_rewards_template(company_id: int, path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Rewards"
    headers = [
        "№ п/п",
        "ФИО",
        "Вид поощрения",
        "Состояние",
        "Указание на вручение",
        "Дата вручения",
        "Примечание",
    ]
    ws.append(headers)

    employments = Employment.query.filter_by(company_id=company_id).all()
    row_num = 1
    for employment in employments:
        rewards = (
            Reward.query.filter_by(employment_id=employment.id)
            .order_by(Reward.updated_at.desc())
            .all()
        )
        person_name = get_current_name(employment.person)
        if not rewards:
            continue
        for reward in rewards:
            ws.append(
                [
                    row_num,
                    person_name,
                    reward.reward_type,
                    STATUS_LABELS.get(reward.status, reward.status),
                    reward.directive_text or "",
                    format_display_date_ru(reward.delivered_date) if reward.delivered_date else "",
                    reward.notes or "",
                ]
            )
            row_num += 1

    if row_num == 1:
        # Keep an example row when company has no rewards yet.
        example_employment = (
            Employment.query.filter_by(company_id=company_id)
            .order_by(Employment.hire_date.desc())
            .first()
        )
        example_name = (
            get_current_name(example_employment.person) if example_employment else "Иванов Иван Иванович"
        )
        ws.append(
            [
                1,
                example_name,
                "Благодарность",
                STATUS_LABELS[RewardStatus.NOT_DELIVERED.value],
                "",
                "",
                "",
            ]
        )

    wb.save(path)
