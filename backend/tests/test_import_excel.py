from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook

from app.extensions import db
from app.models import (
    Company,
    EducationStatus,
    EmployeeGradeHistory,
    Event,
    EventSource,
    GradeCatalog,
    ImportJob,
    ImportStatus,
    Passport,
    Person,
    Role,
    RoleName,
    TenureAward,
    User,
)
from app.services.employees import create_person_with_employment
from app.services.import_excel import (
    annotate_unknown_grades,
    confirm_import,
    dry_run_import,
    parse_workbook,
)
from app.utils.dates import parse_flexible_date


def _write_workbook(
    path: Path,
    *,
    russian_dates: bool = False,
    as_datetime_cells: bool = False,
    position_grade: str = "Мидл",
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(
        [
            "ФИО",
            "Должность",
            "Грейд по должности",
            "Фактический грейд",
            "ВУЗ",
            "Окончание договора",
            "Дата получения текущего грейда",
            "Начало работы",
            "Срок окончания паспорта",
        ]
    )
    head = ["Тестов Тест Тестович", "Инженер", position_grade, "Мидл", "Да"]
    if as_datetime_cells:
        dates = [
            datetime(2026, 12, 1),
            datetime(2024, 3, 1),
            datetime(2021, 1, 10),
            datetime(2029, 8, 20),
        ]
    elif russian_dates:
        dates = [
            "1 декабря 2026 г.",
            "1 марта 2024 г.",
            "10 января 2021 г.",
            "20 августа 2029 г.",
        ]
    else:
        dates = ["01.12.2026", "01.03.2024", "10.01.2021", "20.08.2029"]
    ws.append([*head, *dates])
    wb.save(path)


def _seed_import_company():
    company = Company(name="Import Co")
    role = Role(name=RoleName.HR.value)
    db.session.add_all([company, role])
    db.session.flush()

    user = User(username="importer", full_name="Importer", role_id=role.id)
    user.set_password("secret123")
    db.session.add(user)
    db.session.add_all(
        [
            GradeCatalog(name="Мидл", rank=3, min_years=1.5),
            GradeCatalog(name="Сеньор", rank=4, min_years=2),
        ]
    )
    db.session.commit()
    return company, user


def test_parse_flexible_date_formats():
    assert parse_flexible_date("10.01.2021") == date(2021, 1, 10)
    assert parse_flexible_date("14 ноября 2026 г.") == date(2026, 11, 14)
    assert parse_flexible_date("14 ноября 2026") == date(2026, 11, 14)
    assert parse_flexible_date("2021-01-10 00:00:00") == date(2021, 1, 10)
    assert parse_flexible_date(datetime(2021, 1, 10, 15, 30)) == date(2021, 1, 10)
    assert parse_flexible_date("") is None


def test_confirm_import_runs_rule_engine(app, tmp_path):
    with app.app_context():
        company, user = _seed_import_company()

        path = tmp_path / "employees.xlsx"
        # Actual grade below the position grade, so a grade review is expected.
        _write_workbook(path, position_grade="Сеньор")

        job = ImportJob(
            company_id=company.id,
            filename=path.name,
            uploaded_by_id=user.id,
        )
        db.session.add(job)
        db.session.flush()

        rows = parse_workbook(path)
        dry_run_import(job, rows)
        db.session.commit()

        assert job.status == ImportStatus.VALIDATED.value
        assert job.summary["create"] == 1
        assert "uuid" not in (job.rows[0].raw_data or {})

        confirm_import(job)

        assert job.status == ImportStatus.CONFIRMED.value
        assert job.summary["created"] == 1
        assert job.summary["updated"] == 0
        assert job.rows[0].result == "created"
        events = Event.query.filter_by(
            company_id=company.id,
            source=EventSource.RULE.value,
        ).all()
        event_types = {event.event_type for event in events}
        assert "report" in event_types
        assert "grade" in event_types
        assert "passport" in event_types
        assert len(events) == 3


def test_confirm_import_accepts_datetime_cells(app, tmp_path):
    with app.app_context():
        company, user = _seed_import_company()
        path = tmp_path / "employees-dt.xlsx"
        _write_workbook(path, as_datetime_cells=True)

        job = ImportJob(company_id=company.id, filename=path.name, uploaded_by_id=user.id)
        db.session.add(job)
        db.session.flush()
        dry_run_import(job, parse_workbook(path))
        db.session.commit()

        confirm_import(job)
        assert job.status == ImportStatus.CONFIRMED.value
        assert job.summary["created"] == 1
        assert job.rows[0].result == "created"


def test_confirm_import_accepts_russian_long_dates(app, tmp_path):
    with app.app_context():
        company, user = _seed_import_company()
        path = tmp_path / "employees-ru.xlsx"
        _write_workbook(path, russian_dates=True)

        job = ImportJob(company_id=company.id, filename=path.name, uploaded_by_id=user.id)
        db.session.add(job)
        db.session.flush()
        dry_run_import(job, parse_workbook(path))
        db.session.commit()

        confirm_import(job)
        assert job.status == ImportStatus.CONFIRMED.value
        assert job.summary["created"] == 1


def test_ambiguous_requires_user_choice(app, tmp_path):
    with app.app_context():
        company, user = _seed_import_company()
        create_person_with_employment(
            company_id=company.id,
            full_name="Тестов Тест Тестович",
            hire_date=date(2019, 1, 1),
            title="Старый",
        )
        create_person_with_employment(
            company_id=company.id,
            full_name="Тестов Тест Тестович",
            hire_date=date(2020, 1, 1),
            title="Другой",
        )
        db.session.commit()

        path = tmp_path / "dupes.xlsx"
        _write_workbook(path)
        job = ImportJob(company_id=company.id, filename=path.name, uploaded_by_id=user.id)
        db.session.add(job)
        db.session.flush()
        dry_run_import(job, parse_workbook(path))
        db.session.commit()

        assert job.summary["ambiguous"] == 1
        assert job.rows[0].candidates and len(job.rows[0].candidates) == 2

        people_before = Person.query.count()
        confirm_import(job)
        assert Person.query.count() == people_before
        assert job.summary["skipped"] == 1
        assert job.rows[0].result == "skipped"

        target_uuid = job.rows[0].candidates[0]["uuid"]
        job.status = ImportStatus.VALIDATED.value
        db.session.commit()
        confirm_import(job, {job.rows[0].id: f"update:{target_uuid}"})
        assert job.summary["updated"] == 1
        assert job.rows[0].result == "updated"


def test_reimport_does_not_duplicate_passport_and_grade(app, tmp_path):
    with app.app_context():
        company, user = _seed_import_company()
        path = tmp_path / "once.xlsx"
        _write_workbook(path)

        job1 = ImportJob(company_id=company.id, filename=path.name, uploaded_by_id=user.id)
        db.session.add(job1)
        db.session.flush()
        dry_run_import(job1, parse_workbook(path))
        db.session.commit()
        confirm_import(job1)

        job2 = ImportJob(company_id=company.id, filename=path.name, uploaded_by_id=user.id)
        db.session.add(job2)
        db.session.flush()
        dry_run_import(job2, parse_workbook(path))
        db.session.commit()
        assert job2.summary["update"] == 1
        confirm_import(job2)

        assert Passport.query.count() == 1
        assert EmployeeGradeHistory.query.count() == 1
        assert job2.summary["updated"] == 1


def test_empty_university_does_not_overwrite(app, tmp_path):
    with app.app_context():
        company, user = _seed_import_company()
        person, _employment = create_person_with_employment(
            company_id=company.id,
            full_name="Тестов Тест Тестович",
            hire_date=date(2021, 1, 10),
            title="Инженер",
            education_status=EducationStatus.YES.value,
        )
        db.session.commit()

        wb = Workbook()
        ws = wb.active
        ws.append(["ФИО", "Начало работы", "ВУЗ"])
        ws.append(["Тестов Тест Тестович", "10.01.2021", ""])
        path = tmp_path / "uni.xlsx"
        wb.save(path)

        job = ImportJob(company_id=company.id, filename=path.name, uploaded_by_id=user.id)
        db.session.add(job)
        db.session.flush()
        dry_run_import(job, parse_workbook(path))
        db.session.commit()
        confirm_import(job)

        db.session.refresh(person)
        assert person.education_status == EducationStatus.YES.value


def _write_minimal_workbook(path: Path, full_name: str, hire_date: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(["ФИО", "Начало работы", "ВУЗ"])
    ws.append([full_name, hire_date, "Нет"])
    wb.save(path)


def test_new_employee_import_requires_university_value(app, tmp_path):
    with app.app_context():
        company, user = _seed_import_company()
        path = tmp_path / "missing-university.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.append(["ФИО", "Начало работы", "ВУЗ"])
        ws.append(["Нет Данных", "01.01.2024", ""])
        wb.save(path)

        job = ImportJob(company_id=company.id, filename=path.name, uploaded_by_id=user.id)
        db.session.add(job)
        db.session.flush()
        dry_run_import(job, parse_workbook(path))
        db.session.commit()

        assert job.summary["error"] == 1
        assert "укажите ВУЗ" in job.rows[0].errors[0]


def _tenure_map(employment_id: int) -> dict[int, TenureAward]:
    awards = TenureAward.query.filter_by(employment_id=employment_id).all()
    return {award.milestone_years: award for award in awards}


def test_confirm_import_auto_marks_reached_tenure(app, tmp_path):
    with app.app_context():
        company, user = _seed_import_company()
        path = tmp_path / "tenure-new.xlsx"
        # Hired long ago: 10 and 15 year milestones are already reached, 20 is not.
        _write_minimal_workbook(path, "Стажов Стаж Стажович", "01.01.2010")

        job = ImportJob(company_id=company.id, filename=path.name, uploaded_by_id=user.id)
        db.session.add(job)
        db.session.flush()
        dry_run_import(job, parse_workbook(path))
        db.session.commit()

        confirm_import(job)

        assert job.summary["created"] == 1
        assert job.summary["tenure_marked"] == 2

        person = Person.query.one()
        awards = _tenure_map(person.employments[0].id)
        assert awards[10].is_received is True
        assert awards[10].received_date == date(2020, 1, 1)
        assert awards[15].is_received is True
        assert awards[15].received_date == date(2025, 1, 1)
        assert awards[20].is_received is False
        assert awards[20].received_date is None


def test_confirm_import_can_disable_tenure_marking(app, tmp_path):
    with app.app_context():
        company, user = _seed_import_company()
        path = tmp_path / "tenure-off.xlsx"
        _write_minimal_workbook(path, "Стажов Стаж Стажович", "01.01.2010")

        job = ImportJob(company_id=company.id, filename=path.name, uploaded_by_id=user.id)
        db.session.add(job)
        db.session.flush()
        dry_run_import(job, parse_workbook(path))
        db.session.commit()

        confirm_import(job, mark_reached_tenure=False)

        assert job.summary["tenure_marked"] == 0
        person = Person.query.one()
        awards = _tenure_map(person.employments[0].id)
        assert all(not award.is_received for award in awards.values())


def test_confirm_import_update_existing_tenure_requires_flag(app, tmp_path):
    with app.app_context():
        company, user = _seed_import_company()
        person, _employment = create_person_with_employment(
            company_id=company.id,
            full_name="Стажов Стаж Стажович",
            hire_date=date(2010, 1, 1),
            title="Инженер",
        )
        db.session.commit()

        path = tmp_path / "tenure-update.xlsx"
        _write_minimal_workbook(path, "Стажов Стаж Стажович", "01.01.2010")

        # Update without the flag: existing employee tenure stays untouched.
        job1 = ImportJob(company_id=company.id, filename=path.name, uploaded_by_id=user.id)
        db.session.add(job1)
        db.session.flush()
        dry_run_import(job1, parse_workbook(path))
        db.session.commit()
        assert job1.summary["update"] == 1
        confirm_import(job1, update_existing_tenure=False)

        assert job1.summary["tenure_marked"] == 0
        awards = _tenure_map(person.employments[0].id)
        assert all(not award.is_received for award in awards.values())

        # Default (update existing on): reached milestones get marked.
        job2 = ImportJob(company_id=company.id, filename=path.name, uploaded_by_id=user.id)
        db.session.add(job2)
        db.session.flush()
        dry_run_import(job2, parse_workbook(path))
        db.session.commit()
        confirm_import(job2)

        assert job2.summary["tenure_marked"] == 2
        awards = _tenure_map(person.employments[0].id)
        assert awards[10].is_received is True
        assert awards[15].is_received is True
        assert awards[20].is_received is False


def test_confirm_import_updates_hire_date_and_tenure_milestones(app, tmp_path):
    with app.app_context():
        company, user = _seed_import_company()
        person, employment = create_person_with_employment(
            company_id=company.id,
            full_name="Стажов Стаж Стажович",
            hire_date=date(2020, 1, 1),
            title="Инженер",
        )
        db.session.commit()
        employment_id = employment.id

        path = tmp_path / "tenure-hire-date.xlsx"
        _write_minimal_workbook(path, "Стажов Стаж Стажович", "01.01.2010")

        job = ImportJob(company_id=company.id, filename=path.name, uploaded_by_id=user.id)
        db.session.add(job)
        db.session.flush()
        dry_run_import(job, parse_workbook(path))
        db.session.commit()
        confirm_import(job)

        db.session.refresh(person)
        updated = person.employments[0]
        assert updated.hire_date == date(2010, 1, 1)
        awards = _tenure_map(employment_id)
        assert awards[10].milestone_date == date(2020, 1, 1)
        assert awards[10].is_received is True
        assert awards[15].milestone_date == date(2025, 1, 1)
        assert awards[15].is_received is True
        assert awards[20].is_received is False


def _write_grades_workbook(path: Path, rows: list[list]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(
        [
            "ФИО",
            "Должность",
            "Грейд по должности",
            "Фактический грейд",
            "ВУЗ",
            "Окончание договора",
            "Дата получения текущего грейда",
            "Начало работы",
            "Срок окончания паспорта",
        ]
    )
    for row in rows:
        ws.append(row)
    wb.save(path)


def _run_dry_import(company_id: int, user_id: int, path: Path) -> ImportJob:
    job = ImportJob(company_id=company_id, filename=path.name, uploaded_by_id=user_id)
    db.session.add(job)
    db.session.flush()
    dry_run_import(job, parse_workbook(path))
    db.session.commit()
    return job


def test_dry_run_marks_unknown_grades(app, tmp_path):
    with app.app_context():
        company, user = _seed_import_company()
        path = tmp_path / "unknown-grade.xlsx"
        _write_grades_workbook(
            path,
            [
                [
                    "Новиков Новик Новикович",
                    "Инженер",
                    "Лид",
                    "Лид",
                    "Да",
                    "01.12.2026",
                    "01.03.2024",
                    "10.01.2021",
                    "20.08.2029",
                ]
            ],
        )

        job = _run_dry_import(company.id, user.id, path)

        assert job.summary["create"] == 1
        assert job.summary["unknown_grades"] == [{"name": "Лид", "count": 1}]
        assert job.rows[0].warnings == ["Грейд «Лид» не найден в справочнике"]


def test_unknown_grades_are_deduplicated_across_rows(app, tmp_path):
    with app.app_context():
        company, user = _seed_import_company()
        path = tmp_path / "dup-unknown.xlsx"
        _write_grades_workbook(
            path,
            [
                [
                    "Первый Сотрудник",
                    "Инженер",
                    "Лид",
                    "лид",
                    "Да",
                    "01.12.2026",
                    "01.03.2024",
                    "10.01.2021",
                    "20.08.2029",
                ],
                [
                    "Второй Сотрудник",
                    "Инженер",
                    "Принципал",
                    "Лид",
                    "Нет",
                    "01.12.2026",
                    "01.03.2024",
                    "10.01.2022",
                    "20.08.2029",
                ],
            ],
        )

        job = _run_dry_import(company.id, user.id, path)

        assert job.summary["unknown_grades"] == [
            {"name": "Лид", "count": 2},
            {"name": "Принципал", "count": 1},
        ]
        assert job.rows[0].warnings == ["Грейд «Лид» не найден в справочнике"]
        assert job.rows[1].warnings == [
            "Грейд «Принципал» не найден в справочнике",
            "Грейд «Лид» не найден в справочнике",
        ]


def test_existing_and_inactive_grades_are_not_unknown(app, tmp_path):
    with app.app_context():
        company, user = _seed_import_company()
        db.session.add(GradeCatalog(name="Архив", rank=1, min_years=1, is_active=False))
        db.session.commit()

        path = tmp_path / "known-grades.xlsx"
        _write_grades_workbook(
            path,
            [
                [
                    "Тестов Тест Тестович",
                    "Инженер",
                    "мидл",
                    "Архив",
                    "Да",
                    "01.12.2026",
                    "01.03.2024",
                    "10.01.2021",
                    "20.08.2029",
                ]
            ],
        )

        job = _run_dry_import(company.id, user.id, path)

        assert job.summary["unknown_grades"] == []
        assert job.rows[0].warnings is None


def test_error_rows_do_not_contribute_unknown_grades(app, tmp_path):
    with app.app_context():
        company, user = _seed_import_company()
        path = tmp_path / "error-unknown.xlsx"
        _write_grades_workbook(
            path,
            [
                [
                    "",
                    "Инженер",
                    "Лид",
                    "Лид",
                    "Да",
                    "01.12.2026",
                    "01.03.2024",
                    "10.01.2021",
                    "20.08.2029",
                ]
            ],
        )

        job = _run_dry_import(company.id, user.id, path)

        assert job.summary["error"] == 1
        assert job.summary["unknown_grades"] == []
        assert all(
            not str(item).startswith("Грейд «") for item in (job.rows[0].warnings or [])
        )


def test_revalidate_and_confirm_after_creating_unknown_grade(app, tmp_path):
    with app.app_context():
        company, user = _seed_import_company()
        path = tmp_path / "create-then-import.xlsx"
        _write_grades_workbook(
            path,
            [
                [
                    "Новиков Новик Новикович",
                    "Инженер",
                    "Лид",
                    "Лид",
                    "Да",
                    "01.12.2026",
                    "01.03.2024",
                    "10.01.2021",
                    "20.08.2029",
                ]
            ],
        )

        job = _run_dry_import(company.id, user.id, path)
        assert job.summary["unknown_grades"] == [{"name": "Лид", "count": 1}]

        db.session.add(GradeCatalog(name="Лид", rank=1, min_years=1))
        db.session.commit()

        annotate_unknown_grades(job)
        db.session.commit()

        assert job.summary["unknown_grades"] == []
        assert job.rows[0].warnings is None

        confirm_import(job)

        assert job.status == ImportStatus.CONFIRMED.value
        assert job.summary["created"] == 1
        grade = GradeCatalog.query.filter_by(name="Лид").one()
        history = EmployeeGradeHistory.query.one()
        assert history.grade_id == grade.id


def test_revalidate_api_clears_unknown_grades(admin_client, seed_company, app, tmp_path):
    with app.app_context():
        db.session.add(GradeCatalog(name="Мидл", rank=1, min_years=1.5))
        db.session.commit()
        company_id = seed_company.id

    path = tmp_path / "api-unknown.xlsx"
    _write_grades_workbook(
        path,
        [
            [
                "Новиков Новик Новикович",
                "Инженер",
                "Лид",
                "Лид",
                "Да",
                "01.12.2026",
                "01.03.2024",
                "10.01.2021",
                "20.08.2029",
            ]
        ],
    )

    with path.open("rb") as handle:
        upload = admin_client.post(
            "/api/import/upload",
            data={
                "file": (handle, path.name),
                "company_id": str(company_id),
                "import_type": "employees",
            },
            content_type="multipart/form-data",
        )
    assert upload.status_code == 201
    payload = upload.get_json()["data"]
    assert payload["unknown_grades"] == [{"name": "Лид", "count": 1}]
    assert "unknown_grades" not in (payload["summary"] or {})
    job_id = payload["id"]

    created = admin_client.post(
        "/api/grade-catalog",
        json={"name": "Лид", "rank": 2, "min_years": 2},
    )
    assert created.status_code == 201

    revalidate = admin_client.post(f"/api/import/{job_id}/revalidate", json={})
    assert revalidate.status_code == 200
    data = revalidate.get_json()["data"]
    assert data["unknown_grades"] == []
    assert data["rows"][0]["warnings"] is None

    confirm = admin_client.post(f"/api/import/{job_id}/confirm", json={"row_actions": {}})
    assert confirm.status_code == 200
    assert confirm.get_json()["data"]["status"] == "confirmed"

    with app.app_context():
        grade = GradeCatalog.query.filter_by(name="Лид").one()
        history = EmployeeGradeHistory.query.one()
        assert history.grade_id == grade.id


def test_import_with_contract_term_years(admin_client, seed_company, app, tmp_path):
    with app.app_context():
        db.session.add(GradeCatalog(name="Мидл", rank=1, min_years=1.5))
        db.session.commit()
        company_id = seed_company.id

    path = tmp_path / "contract_term.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(
        [
            "ФИО",
            "Должность",
            "Грейд по должности",
            "Фактический грейд",
            "ВУЗ",
            "Срок договора (лет)",
            "Дата получения текущего грейда",
            "Начало работы",
            "Срок окончания паспорта",
        ]
    )
    ws.append(
        [
            "Сроков Срок Срокович",
            "Инженер",
            "Мидл",
            "Мидл",
            "Да",
            "3",
            "01.03.2024",
            "10.01.2024",
            "20.08.2029",
        ]
    )
    wb.save(path)

    with path.open("rb") as handle:
        upload = admin_client.post(
            "/api/import/upload",
            data={
                "file": (handle, path.name),
                "company_id": str(company_id),
                "import_type": "employees",
            },
            content_type="multipart/form-data",
        )
    assert upload.status_code == 201
    job_id = upload.get_json()["data"]["id"]

    confirm = admin_client.post(f"/api/import/{job_id}/confirm", json={"row_actions": {}})
    assert confirm.status_code == 200

    with app.app_context():
        from app.models import Contract
        contract = Contract.query.first()
        assert contract is not None
        assert contract.term_years == 3.0
        assert contract.end_date == date(2027, 1, 10)


def test_import_with_only_contract_end_calculates_term_years(admin_client, seed_company, app, tmp_path):
    with app.app_context():
        db.session.add(GradeCatalog(name="Мидл", rank=1, min_years=1.5))
        db.session.commit()
        company_id = seed_company.id

    path = tmp_path / "contract_end_only.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(
        [
            "ФИО",
            "Должность",
            "Грейд по должности",
            "Фактический грейд",
            "ВУЗ",
            "Окончание договора",
            "Дата получения текущего грейда",
            "Начало работы",
            "Срок окончания паспорта",
        ]
    )
    ws.append(
        [
            "Авторасчетов Авторасчет",
            "Инженер",
            "Мидл",
            "Мидл",
            "Да",
            "10.01.2026",
            "01.03.2024",
            "10.01.2024",
            "20.08.2029",
        ]
    )
    wb.save(path)

    with path.open("rb") as handle:
        upload = admin_client.post(
            "/api/import/upload",
            data={
                "file": (handle, path.name),
                "company_id": str(company_id),
                "import_type": "employees",
            },
            content_type="multipart/form-data",
        )
    assert upload.status_code == 201
    job_id = upload.get_json()["data"]["id"]

    confirm = admin_client.post(f"/api/import/{job_id}/confirm", json={"row_actions": {}})
    assert confirm.status_code == 200

    with app.app_context():
        from app.models import Contract
        contract = Contract.query.first()
        assert contract is not None
        assert contract.term_years == 2.0
        assert contract.end_date == date(2026, 1, 10)


